# reports/exports/excel.py
import json
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def export_to_excel(data, filename="report.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    if not data:
        # If no data, write a default message
        sheet.cell(row=1, column=1, value="No data available")
    else:
        headers = data[0].keys()

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center")
            # Optional: auto-width
            column_width = max(len(header), 12)
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = column_width

        # Write data rows
        for row_num, row in enumerate(data, start=2):
            for col_num, key in enumerate(headers, start=1):
                value = row.get(key, "")
                sheet.cell(row=row_num, column=col_num, value=value)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

def generate_excel_report(data, filename="report.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    flattened_data = []

    # Flatten if there are nested lists like 'sales' or 'purchases'
    for row in data:
        base = {k: v for k, v in row.items() if not isinstance(v, list)}

        for key in ['sales', 'purchases']:
            if key in row and isinstance(row[key], list):
                for entry in row[key]:
                    if isinstance(entry, dict):
                        new_row = base.copy()
                        for field_key, field_val in entry.items():
                            new_row[f"{key}_{field_key}"] = field_val
                        flattened_data.append(new_row)
            else:
                flattened_data.append(base)

    if not flattened_data:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Report"
        sheet.cell(row=1, column=1, value="No data available")

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    headers = list(flattened_data[0].keys())

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_num, row in enumerate(flattened_data, start=2):
        for col_num, key in enumerate(headers, start=1):
            value = row.get(key, '')
            if isinstance(value, (dict, list)):
                value = json.dumps(value)  # convert complex types to JSON string
            sheet.cell(row=row_num, column=col_num, value=value)

    # Return as HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
