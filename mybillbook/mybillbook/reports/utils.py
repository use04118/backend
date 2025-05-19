from django.http import HttpResponse
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from functools import wraps
from django.http import JsonResponse
import json
from .export import generate_excel_report

def generate_excel(headers, data, filename="report.xlsx"):
    """
    Generates an Excel file from the provided headers and data.
    """
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Add headers to the sheet
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Add data to the sheet
    for row_num, row in enumerate(data, 2):
        for col_num, value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = value

    # Prepare the response for Excel download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"

    # Save the workbook into the response
    wb.save(response)

    return response

# utils/decorators.py

def export_excel_view(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        # Call the original view
        response = view_func(request, *args, **kwargs)

        # Check for ?export_excel=true
        if request.GET.get('export_excel') != 'true':
            return response

        try:
            # Parse response content (works for JsonResponse or DRF Response)
            if hasattr(response, 'data'):
                data = response.data
            else:
                # Handle JsonResponse: decode JSON from .content
                data = json.loads(response.content)

            # Normalize data to list of dicts for Excel
            if isinstance(data, dict) and 'transactions' in data:
                records = data['transactions']
            elif isinstance(data, dict):
                records = []
                for k, v in data.items():
                    if isinstance(v, dict):
                        record = {"Party Name": k, **v}
                    else:
                        record = {"Party Name": k, "Value": v}
                    records.append(record)
            elif isinstance(data, list):
                records = data
            else:
                records = [{"data": str(data)}]

            return generate_excel_report(records, filename="report.xlsx")

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return wrapper
