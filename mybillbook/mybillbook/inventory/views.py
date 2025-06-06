from django.shortcuts import get_object_or_404
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
from django.db.models import F
from django.http import JsonResponse
from django_filters.rest_framework import DjangoFilterBackend
from .filter import ItemFilter
from .models import Item, ItemCategory, MeasuringUnit, GSTTaxRate, Service
from .serializers import (
    ItemSerializer, ItemCategorySerializer, MeasuringUnitSerializer,
    GSTTaxRateSerializer, ServiceSerializer
)
from users.utils import get_current_business, log_action
from .permissions import HasItemPermission
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from rest_framework.parsers import MultiPartParser
from django.db import transaction
from datetime import date
from users.models import Business
from godown.models import Godown
from decimal import Decimal



@api_view(['GET'])
@permission_classes([IsAuthenticated, HasItemPermission])
def stock_value(request):
    business = get_current_business(request.user) 
    # Query all items where itemType is 'Product'
   
    items = Item.objects.filter(business=business, itemType='Product')


    stock_values = []
    total_stock_value = 0  # Initialize the total stock value

    for item in items:
        # Calculate purchase price without tax
        def get_purchasePrice_without_tax(item):
            """Returns the tax-inclusive purchase price if stored without tax."""
            if item.purchasePriceType == "With Tax":
                return item.calculate_price(item.purchasePrice, "With Tax")
            return item.purchasePrice  # Already tax-inclusive
        purchase_price_without_tax = get_purchasePrice_without_tax(item)
        
        # Calculate stock value based on closingStock and purchasePrice_without_tax
        if item.closingStock and purchase_price_without_tax:
            stock_value = item.closingStock * purchase_price_without_tax
        else:
            stock_value = 0.00
        
        # Add the stock value to the total stock value
        total_stock_value += stock_value
        
        # Store the result in a dictionary to return as JSON or log
        stock_values.append({
            "item_name": item.itemName,
            "stock_value": round(stock_value,2)
        })

    # Return as a JSON response with total stock value included
    return JsonResponse({
        "stock_values": stock_values,
        "total_stock_value": round(total_stock_value,2)
    })


# Low Stock Count
@api_view(['GET'])
@permission_classes([IsAuthenticated, HasItemPermission])
def low_stock_value(request):
    low_stock_count = Item.objects.filter(enableLowStockWarning=True).count()
    return Response({'totalLowStockItems': low_stock_count})


# Item CRUD Operations
class ItemListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated, HasItemPermission]
    serializer_class = ItemSerializer
    filter_backends = [DjangoFilterBackend]

    def get_queryset(self):
        return Item.objects.filter(business=get_current_business(self.request.user))
    
    def perform_create(self, serializer):
        business = get_current_business(self.request.user)
        serializer.save(business=business)
        # data= request.data
        # many = isinstance(data, list)
        # serializer = self.get_serializer(data=data, many=many)
        instances = serializer.instance
        items = instances if isinstance(instances, list) else [instances]
       

        for item in items:
            if item.itemType == "Product":
                if item.closingStock is not None and item.lowStockQty is not None:
                    item.enableLowStockWarning = item.closingStock <= item.lowStockQty
                else:
                    item.enableLowStockWarning = False  # Default to False if None

            elif item.itemType == "Service":
                # Handle Service-specific fields
                item.closingStock = None
                item.lowStockQty = None
                item.enableLowStockWarning = None

            item.save()

            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

class ItemDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ItemSerializer
    permission_classes = [IsAuthenticated, HasItemPermission]

    
    def get_queryset(self):
        return Item.objects.filter(business=get_current_business(self.request.user))

    def perform_update(self, serializer):
        # Save the item and handle service/product-specific updates
        item = serializer.save()
        if item.itemType == "Product":
            if item.closingStock is not None and item.lowStockQty is not None:
                item.enableLowStockWarning = item.closingStock <= item.lowStockQty
            else:
                item.enableLowStockWarning = False  # Default to False if None
        
        elif item.itemType == "Service":
            # Handle Service-specific fields
            item.closingStock = None
            item.lowStockQty = None
            item.enableLowStockWarning = None
        
        item.save()

        log_action(self.request.user,get_current_business(self.request.user),"item_updated",{"name": item.itemName} ) # adjust the field name as needed)

    def perform_destroy(self, instance):
        log_action(self.request.user, get_current_business(self.request.user), "item_deleted", {"name": instance.itemName})
        instance.delete()


class ServiceListCreateView(generics.ListCreateAPIView):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Service.objects.filter(business=get_current_business(self.request.user))

    def perform_create(self, serializer):
        business = get_current_business(self.request.user)
        serializer.save(business=business)


class ServiceDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated, HasItemPermission]

    
    def get_queryset(self):
        return Service.objects.filter(business=get_current_business(self.request.user))

    def perform_update(self, serializer):
        # Save the service
        service = serializer.save()
        # You can add additional logic for service if needed


class ItemListView(generics.ListAPIView):
    filterset_class = ItemFilter


# List and Create API for Categories
class ItemCategoryListCreateView(generics.ListCreateAPIView):
    serializer_class = ItemCategorySerializer
    permission_classes = [IsAuthenticated, HasItemPermission]

    def get_queryset(self):
        return ItemCategory.objects.filter(business=get_current_business(self.request.user))

    def create(self, request, *args, **kwargs):
        business = get_current_business(request.user)
        data = request.data
        many = isinstance(data, list)
        serializer = self.get_serializer(data=data, many=many)
        if serializer.is_valid():
            serializer.save(business=business)
            if many:
                for cat in serializer.instance:
                    log_action(request.user, business, "item_category_created", {"name": cat.name})
            else:
                log_action(request.user, business, "item_category_created", {"name": serializer.instance.name})
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

# Retrieve, Update, Destroy API for Categories
class ItemCategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ItemCategorySerializer
    permission_classes = [IsAuthenticated, HasItemPermission]

    
    def get_queryset(self):
        return ItemCategory.objects.filter(business=get_current_business(self.request.user))

    def perform_update(self, serializer):
        category = serializer.save()
        log_action(self.request.user, get_current_business(self.request.user), "item_category_updated", {"name": category.name})

    def perform_destroy(self, instance):
        log_action(self.request.user, get_current_business(self.request.user), "item_category_deleted", {"name": instance.name})
        instance.delete()



class MeasuringUnitListCreateView(generics.ListCreateAPIView):
    serializer_class = MeasuringUnitSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return MeasuringUnit.objects.all()

    def create(self, request, *args, **kwargs):
        data = request.data
        many = isinstance(data, list)
        serializer = self.get_serializer(data=data, many=many)
        if serializer.is_valid():
            serializer.save()  # No need to attach business
            if many:
                for unit in serializer.instance:
                    log_action(request.user, None, "measuringunit_created", {"name": unit.name})
            else:
                log_action(request.user, None, "measuringunit_created", {"name": serializer.instance.name})
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


class MeasuringUnitDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = MeasuringUnitSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return MeasuringUnit.objects.all()


# --- GST Tax Rates ---

class GSTTaxRateListCreateView(generics.ListCreateAPIView):
    serializer_class = GSTTaxRateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return GSTTaxRate.objects.all()

    def create(self, request, *args, **kwargs):
        data = request.data
        many = isinstance(data, list)
        serializer = self.get_serializer(data=data, many=many)
        if serializer.is_valid():
            serializer.save()  # No need to attach business
            if many:
                for gst in serializer.instance:
                    log_action(request.user, None, "gsttaxrate_created", {
                        "rate": str(gst.rate),
                        "description": gst.description
                    })
            else:
                gst = serializer.instance
                log_action(request.user, None, "gsttaxrate_created", {
                    "rate": str(gst.rate),
                    "description": gst.description
                })
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


class GSTTaxRateDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = GSTTaxRateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return GSTTaxRate.objects.all()

    def perform_update(self, serializer):
        gst = serializer.save()
        log_action(self.request.user, None, "gsttaxrate_updated", {
            "rate": str(gst.rate),
            "description": gst.description
        })

    def perform_destroy(self, instance):
        log_action(self.request.user, None, "gsttaxrate_deleted", {
            "rate": str(instance.rate),
            "description": instance.description
        })
        instance.delete()

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def bulk_update_items_from_excel(request):
    user = request.user
    business = user.current_business

    if not business:
        return Response({"error": "No active business selected for this user."}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = openpyxl.load_workbook(file)
        bulk_sheet = wb["bulk_upload"]
        ref_sheet = wb["ReferenceData"]

        # Reference mapping
        reference_data = {
            "GST Tax Rate ID": {},
            "Measuring Unit ID": {},
            "Godown ID": {},
        }

            # Reference mapping from fixed columns
        for row in ref_sheet.iter_rows(min_row=2, values_only=True):
            gst_id, gst_desc, unit_id, unit_name, godown_id, godown_name = row[:6]

            # Map GST
            if gst_desc and gst_id:
                reference_data["GST Tax Rate ID"][str(gst_desc).strip()] = gst_id

            # Map Measuring Unit
            if unit_name and unit_id:
                reference_data["Measuring Unit ID"][str(unit_name).strip()] = unit_id

            # Map Godown
            if godown_name and godown_id:
                reference_data["Godown ID"][str(godown_name).strip()] = godown_id
        print("GST Ref Map:", reference_data["GST Tax Rate ID"])



        headers = [cell.value for cell in bulk_sheet[1]]

        created_count = 0
        stock_updated_count = 0
        skipped_count = 0

        with transaction.atomic():
            for row in bulk_sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                item_code = str(row_data.get("Item Code")).strip() if row_data.get("Item Code") else None
                item_name = str(row_data.get("Item Name")).strip() if row_data.get("Item Name") else None
                opening_stock = float(row_data.get("Opening Stock") or 0)

                if not item_code or not item_name:
                    skipped_count += 1
                    continue

                existing_item = Item.objects.filter(
                    business=business,
                    itemCode=item_code,
                    itemName__iexact=item_name
                ).first()

                if existing_item:
                    existing_item.openingStock += Decimal(str(opening_stock))
                    existing_item.closingStock = existing_item.openingStock
                    existing_item.save()
                    stock_updated_count += 1
                    continue

                new_item = Item(
                    business=business,
                    itemCode=item_code,
                    itemName=item_name,
                    openingStock=Decimal(str(opening_stock)),
                    closingStock=Decimal(str(opening_stock)),
                    salesPrice=row_data.get("Sales Price") or 0,
                    purchasePrice=row_data.get("Purchase Price") or 0,
                    salesPriceType=row_data.get("Sales Price Type") or "With Tax",
                    purchasePriceType=row_data.get("Purchase Price Type") or "With Tax",
                    date=row_data.get("Date") or date.today(),
                    itemBatch=row_data.get("Item Batch"),
                    hsnCode=row_data.get("HSN Code"),
                    description=row_data.get("Description"),
                    lowStockQty=row_data.get("Low Stock Qty") or 0,
                )

                # Boolean flag
                low_stock_flag = row_data.get("Enable Low Stock Warning")
                if isinstance(low_stock_flag, str):
                    new_item.enableLowStockWarning = low_stock_flag.strip().lower() in ["true", "yes", "1"]
                elif isinstance(low_stock_flag, bool):
                    new_item.enableLowStockWarning = low_stock_flag
                else:
                    new_item.enableLowStockWarning = False

                # Resolve names to IDs from reference data
                gst_name = row_data.get("GST Description")
                if gst_name:
                    gst_id = reference_data["GST Tax Rate ID"].get(str(gst_name).strip())
                    if not gst_id:
                        return Response({"error": f"Invalid GST Description: '{gst_name}'"}, status=status.HTTP_400_BAD_REQUEST)
                    new_item.gstTaxRate_id = gst_id

                unit_name = row_data.get("Unit")
                if unit_name:
                    unit_id = reference_data["Measuring Unit ID"].get(str(unit_name).strip())
                    if not unit_id:
                        return Response({"error": f"Invalid Measuring Unit: '{unit_name}'"}, status=status.HTTP_400_BAD_REQUEST)
                    new_item.measuringUnit_id = unit_id

                godown_name = row_data.get("Godown")
                if godown_name:
                    godown_id = reference_data["Godown ID"].get(str(godown_name).strip())
                    if not godown_id:
                        return Response({"error": f"Invalid Godown: '{godown_name}'"}, status=status.HTTP_400_BAD_REQUEST)
                    new_item.godown_id = godown_id

                category_name = str(row_data.get("Category Name")).strip() if row_data.get("Category Name") else None
                if not category_name:
                    return Response({"error": f"Category Name is required."}, status=status.HTTP_400_BAD_REQUEST)

                # Try to fetch from DB by name (case-insensitive match)
                category_obj = ItemCategory.objects.filter(name__iexact=category_name, business=business).first()

                if not category_obj:
                    # Create new category if it doesn't exist
                    category_obj = ItemCategory.objects.create(name=category_name, business=business)

                new_item.category = category_obj

                print("new item bulk",new_item)

                new_item.save()
                created_count += 1

        return Response({
            "message": f"{stock_updated_count} items updated, {created_count} created, {skipped_count} skipped."
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

